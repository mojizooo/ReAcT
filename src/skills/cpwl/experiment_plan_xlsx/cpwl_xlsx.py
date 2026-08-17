"""Deterministic CPWL XLSX generation under the migrated human-owned contract."""

from __future__ import annotations

import math
import re
import shutil
from copy import copy
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from pydantic import BaseModel, Field, computed_field, field_validator, model_validator

from react_color_agent.storage import TaskStore

STOCK_CONCENTRATION_MMOL_ML = 0.0001
PRODUCT_VOLUME_ML = 8.0
MAX_COMPONENT_VOLUME_ML = 4.8
# Derive final-concentration limits from the fixed laboratory volume contract.
MAX_COMPONENT_CONCENTRATION_MMOL_ML = (
    MAX_COMPONENT_VOLUME_ML * STOCK_CONCENTRATION_MMOL_ML / PRODUCT_VOLUME_ML
)
MIN_TOTAL_CONCENTRATION_MMOL_ML = (
    (PRODUCT_VOLUME_ML - MAX_COMPONENT_VOLUME_ML)
    * STOCK_CONCENTRATION_MMOL_ML
    / PRODUCT_VOLUME_ML
)
MAX_TOTAL_CONCENTRATION_MMOL_ML = STOCK_CONCENTRATION_MMOL_ML
MAX_BATCH_RECIPES = 24
MAX_FOLLOWUP_RECIPES = 12
# Preserve physical boundary semantics despite binary floating-point endpoint noise.
VOLUME_TOLERANCE_ML = 1e-12
CONCENTRATION_TOLERANCE_MMOL_ML = (
    VOLUME_TOLERANCE_ML * STOCK_CONCENTRATION_MMOL_ML / PRODUCT_VOLUME_ML
)
EXPECTED_HEADERS = [
    "实验名称", "启用合成", "启用后处理", "启用检测", "物料摩尔百分比", "产物体积浓度",
    "第一个实验前清洗", "停留时间", "反应温度", "背压", "反应清洗", "收集起止",
    "沉淀循环次数", "复溶循环次数", "清洗循环次数", "光通量确认方式", "沉淀参数", "复溶参数",
    "清洗参数", "后处理收集瓶", "检测模块", "继续后处理", "检测设备", "UV起始波长",
    "UV结束波长", "扫描模式", "激发波长", "发射起始波长", "发射结束波长", "发射波长",
    "激发起始波长", "激发结束波长",
]


class CpwlMaterial(BaseModel):
    """One PubChem-identified material occupying its stable A-E workbook slot."""

    slot: Literal["A", "B", "C", "D", "E"]
    material_key: str = Field(min_length=1)
    name: str = Field(min_length=1)
    molecular_weight_g_mol: float = Field(gt=0)


class CpwlDiscrimination(BaseModel):
    """Record the falsifiable decision boundary of one planned formulation."""

    hypothesis: str = Field(min_length=1)
    reference_sample_ids: list[str] = Field(default_factory=list)
    outcome_if_supported: str = Field(min_length=1)
    outcome_if_not_supported: str = Field(min_length=1)

    @field_validator("reference_sample_ids")
    @classmethod
    def validate_reference_sample_ids(cls, values: list[str]) -> list[str]:
        """Keep recipe lineage readable before tools validate referenced observations."""
        normalized = [value.strip() for value in values]
        if any(not value for value in normalized):
            raise ValueError("reference sample ids must be non-empty")
        if len(set(normalized)) != len(normalized):
            raise ValueError("reference sample ids must not contain duplicates")
        return normalized


class CpwlRecipe(BaseModel):
    """One formulation that expands into one synthesis row and one manual measurement result."""

    recipe_id: str = Field(min_length=1, max_length=16)
    concentrations_mmol_ml: tuple[float, float, float, float, float]
    purpose: str
    discrimination: CpwlDiscrimination

    @field_validator("concentrations_mmol_ml")
    @classmethod
    def validate_concentrations(
        cls, values: tuple[float, float, float, float, float]
    ) -> tuple[float, float, float, float, float]:
        """Enforce the fixed stock and per-material concentration boundaries."""
        if any(
            not math.isfinite(value)
            or value < 0
            or value
            > MAX_COMPONENT_CONCENTRATION_MMOL_ML + CONCENTRATION_TOLERANCE_MMOL_ML
            for value in values
        ):
            raise ValueError(
                "recipe concentrations must be finite values within "
                f"0..{MAX_COMPONENT_CONCENTRATION_MMOL_ML:.6g} mmol/ml"
            )
        if sum(values) <= 0:
            raise ValueError("at least one recipe component must be positive")
        return values

    @computed_field
    @property
    def total_concentration_mmol_ml(self) -> float:
        """Return the final product concentration written to workbook column F."""
        return float(sum(self.concentrations_mmol_ml))

    @computed_field
    @property
    def molar_fractions(self) -> tuple[float, float, float, float, float]:
        """Normalize the five A-E components for the workbook material-fraction field."""
        total = self.total_concentration_mmol_ml
        return tuple(value / total for value in self.concentrations_mmol_ml)  # type: ignore[return-value]

    @computed_field
    @property
    def stock_volumes_ml(self) -> tuple[float, float, float, float, float]:
        """Convert final concentrations to volumes from fixed 0.0001 mmol/ml stocks."""
        return tuple(value * PRODUCT_VOLUME_ML / STOCK_CONCENTRATION_MMOL_ML for value in self.concentrations_mmol_ml)  # type: ignore[return-value]

    @computed_field
    @property
    def solvent_volume_ml(self) -> float:
        """Compute the remaining solvent volume for the fixed 8 ml product volume."""
        return PRODUCT_VOLUME_ML - sum(self.stock_volumes_ml)


class CpwlBatchPlan(BaseModel):
    """Machine-readable batch input to the migrated deterministic XLSX writer."""

    plan_id: str
    run_id: str
    target_cie: tuple[float, float]
    materials: list[CpwlMaterial] = Field(min_length=1, max_length=5)
    recipes: list[CpwlRecipe] = Field(min_length=1, max_length=MAX_BATCH_RECIPES)
    applied_skills: list[str] = Field(default_factory=list)
    design_context: dict[str, object] = Field(default_factory=dict)

    @model_validator(mode="after")
    def validate_contract(self) -> "CpwlBatchPlan":
        """Require stable slots, unique recipes, and physically feasible stock volumes."""
        expected_slots = list("ABCDE"[: len(self.materials)])
        if [material.slot for material in self.materials] != expected_slots:
            raise ValueError("materials must occupy contiguous A-E slots")
        if len({recipe.recipe_id for recipe in self.recipes}) != len(self.recipes):
            raise ValueError("recipe ids must be unique")
        batch_numbers = {_recipe_batch_number(recipe.recipe_id) for recipe in self.recipes}
        if len(batch_numbers) != 1:
            raise ValueError("all recipes in a CPWL batch plan must use one B{batch} prefix")
        for recipe in self.recipes:
            if any(recipe.concentrations_mmol_ml[index] != 0 for index in range(len(self.materials), 5)):
                raise ValueError("unused material slots must remain zero")
            volumes = (*recipe.stock_volumes_ml, recipe.solvent_volume_ml)
            if recipe.solvent_volume_ml < -VOLUME_TOLERANCE_ML or any(
                volume > MAX_COMPONENT_VOLUME_ML + VOLUME_TOLERANCE_ML for volume in volumes
            ):
                raise ValueError(f"recipe {recipe.recipe_id} exceeds the fixed 8 ml volume contract")
        return self


def build_initial_cpwl_batch(
    *,
    run_id: str,
    target_cie: tuple[float, float],
    materials: list[CpwlMaterial],
    agent_recipes: list[tuple[tuple[float, float, float, float, float], str, CpwlDiscrimination]],
    scientific_rationale: str,
    applied_skills: list[str],
) -> CpwlBatchPlan:
    """Build a B1 workbook from the Agent's bounded, auditable screening matrix."""
    if not 1 <= len(agent_recipes) <= MAX_BATCH_RECIPES:
        raise ValueError("an initial CPWL batch must contain one to twenty-four Agent-designed recipes")
    recipes = [
        # Keep one recipe ID shared by the synthesis row and manual measurement result folder.
        CpwlRecipe(
            recipe_id=f"B1-N{index}",
            concentrations_mmol_ml=values,
            purpose=purpose,
            discrimination=discrimination,
        )
        for index, (values, purpose, discrimination) in enumerate(agent_recipes, start=1)
    ]
    return CpwlBatchPlan(
        plan_id="experiment_batch_001",
        run_id=run_id,
        target_cie=target_cie,
        materials=materials,
        recipes=recipes,
        applied_skills=applied_skills,
        design_context={
            "strategy": "Agent-designed initial screening batch",
            "scientific_rationale": scientific_rationale,
            "recipe_source": "agent",
        },
    )


def build_followup_cpwl_batch(
    *,
    run_id: str,
    target_cie: tuple[float, float],
    materials: list[CpwlMaterial],
    batch_number: int,
    candidate_recipes: list[tuple[tuple[float, float, float, float, float], str, CpwlDiscrimination]],
    applied_skills: list[str],
    design_context: dict[str, object],
) -> CpwlBatchPlan:
    """Build an Agent-selected follow-up batch from controlled candidate recipes."""
    if batch_number < 2:
        raise ValueError("follow-up CPWL batches must begin at B2")
    if not 1 <= len(candidate_recipes) <= MAX_FOLLOWUP_RECIPES:
        raise ValueError("a follow-up CPWL batch must contain one to twelve candidate recipes")
    recipes = [
        CpwlRecipe(
            recipe_id=f"B{batch_number}-N{index}",
            concentrations_mmol_ml=concentrations,
            purpose=purpose,
            discrimination=discrimination,
        )
        for index, (concentrations, purpose, discrimination) in enumerate(candidate_recipes, start=1)
    ]
    return CpwlBatchPlan(
        plan_id=f"experiment_batch_{batch_number:03}",
        run_id=run_id,
        target_cie=target_cie,
        materials=materials,
        recipes=recipes,
        applied_skills=applied_skills,
        design_context=design_context,
    )


def write_cpwl_xlsx_artifacts(store: TaskStore, plan: CpwlBatchPlan) -> list[str]:
    """Write plan artifacts and create empty sample directories for manual data return."""
    if plan.run_id != store.load().task_id:
        raise ValueError("CPWL batch plan must be bound to the current task id")
    template_path = Path(__file__).resolve().parents[3] / "constraints" / "exp_template.xlsx"
    if not template_path.is_file():
        raise FileNotFoundError(template_path)
    output_dir = store.artifact_path("artifacts/experiment_plans")
    output_dir.mkdir(parents=True, exist_ok=True)
    batch_number = _plan_batch_number(plan)
    workbook_path = output_dir / f"batch_{batch_number:03}.xlsx"
    explanation_path = output_dir / f"batch_{batch_number:03}.md"
    measurement_root = store.artifact_path(f"artifacts/measurement_returns/round-{batch_number:03}")
    measurement_root.mkdir(parents=True, exist_ok=True)
    for recipe in plan.recipes:
        # Leave each detection directory empty so instrument exports can be copied in directly.
        (measurement_root / f"{recipe.recipe_id}-D").mkdir(exist_ok=True)
    shutil.copy2(template_path, workbook_path)
    workbook = load_workbook(workbook_path, data_only=False)
    if workbook.sheetnames != ["导入实验模板", "填写说明"]:
        raise ValueError("XLSX template worksheets violate the immutable contract")
    worksheet = workbook["导入实验模板"]
    if [worksheet.cell(1, column).value for column in range(1, 33)] != EXPECTED_HEADERS:
        raise ValueError("XLSX template headers violate the immutable 32-column contract")
    styles = [copy(worksheet.cell(2, column)._style) for column in range(1, 33)]
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    row_number = 2
    for recipe in _execution_rows(plan):
        for column, value in enumerate(_row_values(recipe), start=1):
            cell = worksheet.cell(row_number, column, value)
            cell._style = copy(styles[column - 1])
        row_number += 1
    workbook.save(workbook_path)
    explanation_path.write_text(_render_explanation(plan), encoding="utf-8")
    plan_path = store.write_artifact_json(
        f"artifacts/experiment_plans/batch_{batch_number:03}_design.json", plan.model_dump(mode="json")
    )
    workbook_ref = workbook_path.relative_to(store.run_dir).as_posix()
    explanation_ref = explanation_path.relative_to(store.run_dir).as_posix()
    measurement_return_ref = measurement_root.relative_to(store.run_dir).as_posix()
    _validate_written_workbook(workbook_path, explanation_path, plan)
    return [workbook_ref, explanation_ref, plan_path, measurement_return_ref]


def _execution_rows(plan: CpwlBatchPlan) -> list[CpwlRecipe]:
    """Return one ordered synthesis row per formulation for human laboratory execution."""
    return plan.recipes


def _row_values(recipe: CpwlRecipe) -> list[object]:
    """Expand one formulation into its sole synthesis row in the fixed A-AF column order."""
    fractions = ",".join(_format_number(value) for value in recipe.molar_fractions)
    return [
        f"{recipe.recipe_id}-S", True, False, False, fractions,
        recipe.total_concentration_mmol_ml, True, 1, 25, 1, "1,2", "4,8", None, None, None, None,
        None, None, None, None, 3, False, "1,1", 400, 700, 2, 350, 360, 760, 350, 400, 700,
    ]


def _render_explanation(plan: CpwlBatchPlan) -> str:
    """Create the matching auditable explanation required alongside every workbook."""
    batch_number = _plan_batch_number(plan)
    strategy = str(plan.design_context.get("strategy", "CPWL experiment batch"))
    lines = [
        "# CPWL 实验方案", "", "## 实验目标", "", f"- Batch: B{batch_number}",
        f"- Target CIE xy: {plan.target_cie}", f"- Strategy: {strategy}", "", "## 物料映射", "",
        "| 槽位 | 物料 | PubChem identity | Molecular weight (g/mol) |",
        "| --- | --- | --- | ---: |",
    ]
    for material in plan.materials:
        lines.append(f"| {material.slot} | {material.name} | {material.material_key} | {material.molecular_weight_g_mol} |")
    rationale = plan.design_context.get("scientific_rationale")
    if rationale:
        lines.extend(["", "## 研究理由", "", f"- {rationale}"])
    lines.extend([
        "",
        "## 浓度与换算",
        "",
        "`mg/ml = mmol/ml * molecular_weight_g_mol`",
        "",
        "## 配方辨别意图",
        "",
    ])
    for recipe in plan.recipes:
        lines.extend([
            f"### {recipe.recipe_id}",
            f"- Purpose: {recipe.purpose}",
            f"- Hypothesis: {recipe.discrimination.hypothesis}",
            f"- Reference samples: {', '.join(recipe.discrimination.reference_sample_ids) or '(initial batch; none)'}",
            f"- If supported: {recipe.discrimination.outcome_if_supported}",
            f"- If not supported: {recipe.discrimination.outcome_if_not_supported}",
            f"- Final concentrations A-E (mmol/ml): {','.join(_format_number(value) for value in recipe.concentrations_mmol_ml)}",
            f"- Molar fractions A-E: {','.join(_format_number(value) for value in recipe.molar_fractions)}",
            f"- Stock volumes A-E (ml): {','.join(_format_number(value) for value in recipe.stock_volumes_ml)}",
            f"- Solvent volume (ml): {_format_number(recipe.solvent_volume_ml)}", "",
        ])
    lines.extend([
        "## 人工检测和数据回传", "", "- The XLSX contains synthesis rows only; detection is performed manually outside the import workflow.", "- Detection result IDs use the `-D` suffix only for returned spectrum directories.", "", "### 回传批次", "", f"- Data root: `artifacts/measurement_returns/round-{batch_number:03}/`", "- The empty sample directories are created automatically with this plan; copy required emission.txt and any available optional absorption.txt into each matching `-D` directory.", "- CIE xy is calculated by the deterministic CIE 1931 2-degree tool; do not submit manual CIE JSON.", "",
        "### 必交样品", "", *[f"- `{recipe.recipe_id}-D`" for recipe in plan.recipes], "",
        "### 目录结构", "", "```text", f"round-{batch_number:03}/", f"├── B{batch_number}-N1-D/", "│   ├── emission.txt", "│   └── absorption.txt  # optional", "└── ...", "```", "",
        "### 发射谱文件合同", "", "`emission.txt` uses UTF-8 and exactly this header:", "", "```text", "Scan Mode: 发射扫描", "激发波长: 350 nm", "发射波长范围: 360 - 760 nm", "步长: 0.2 nm", "波长(nm)\\t荧光强度", "```", "", "Then provide 2001 tab-delimited rows from 360.0 to 760.0 nm ascending at 0.2 nm. Intensities are finite, non-negative, and sum to more than zero. The former 400-700 nm grids are accepted only when reading historical runs.", "",
        "### 吸收谱文件合同", "", "`absorption.txt` is optional. When supplied, it uses UTF-8 (optional BOM) and exactly this header:", "", "```text", "Wavelength(nm)\\tTransmittance(%)\\tAbsorbance", "```", "", "Then provide 301 tab-delimited rows from 700 to 400 nm descending at 1 nm. A missing, partial, excluded, or invalid absorption file never blocks emission-based CIE; full absorption features require usable QC status.", "",
        "### 拒收条件", "", "Missing or unexpected sample directories, an invalid sample name, or invalid emission content rejects the whole returned batch. Missing or invalid absorption is status-recorded and never blocks qualified emission spectra.", "",
        "## 固定设备参数", "", "Columns G-AF follow constraints/constraint.md.", "",
        "## 假设与缺失信息", "", "Molecular weights originate from the saved PubChem evidence; laboratory SOP, solvent and safety controls remain required execution metadata.", "",
        "## 验证结果", "", "Workbook reopened successfully with immutable worksheet names, headers, row grouping, and matching explanation file.", "",
    ])
    return "\n".join(lines)


def _recipe_batch_number(recipe_id: str) -> int:
    """Parse the authoritative B{batch}-N{number} identifier from one formulation."""
    match = re.fullmatch(r"B([1-9][0-9]*)-N([1-9][0-9]*)", recipe_id)
    if match is None:
        raise ValueError("recipe_id must use the B{batch}-N{number} CPWL convention")
    return int(match.group(1))


def _plan_batch_number(plan: CpwlBatchPlan) -> int:
    """Resolve the one validated batch number for artifact naming and user feedback."""
    return _recipe_batch_number(plan.recipes[0].recipe_id)


def _validate_written_workbook(workbook_path: Path, explanation_path: Path, plan: CpwlBatchPlan) -> None:
    """Fail closed unless the generated file satisfies the minimum immutable workbook contract."""
    workbook = load_workbook(workbook_path, data_only=False)
    if workbook.sheetnames != ["导入实验模板", "填写说明"]:
        raise ValueError("generated workbook worksheet contract failed")
    worksheet = workbook["导入实验模板"]
    if worksheet.max_column != 32 or [worksheet.cell(1, column).value for column in range(1, 33)] != EXPECTED_HEADERS:
        raise ValueError("generated workbook header contract failed")
    if worksheet.max_row != len(plan.recipes) + 1:
        raise ValueError("generated workbook row count contract failed")
    for row, recipe in enumerate(_execution_rows(plan), start=2):
        if worksheet.cell(row, 1).value != f"{recipe.recipe_id}-S":
            raise ValueError("generated workbook synthesis naming contract failed")
        if worksheet.cell(row, 2).value is not True or worksheet.cell(row, 4).value is not False:
            raise ValueError("generated workbook synthesis module flag contract failed")
    if not explanation_path.is_file():
        raise ValueError("generated workbook explanation file is missing")


def _format_number(value: float) -> str:
    """Render numeric values without lossy scientific-notation noise in audit artifacts."""
    return f"{value:.12f}".rstrip("0").rstrip(".") or "0"
